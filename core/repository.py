from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
import sqlite3
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from . import dataset_repository, db
from .normalization import clean_dataframe
from .settings import (
    CEP_IBGE_DIR,
    CONTRATOS_NEGOCIACOES_DIR,
    CONTRATOS_VIGENTES_DIR,
    LEGACY_DATA_DIR,
    PEDIDOS_DIR,
    REGIOES_LOGISTICAS_DIR,
    SUPPORTED_EXTENSIONS,
)


DATASET_DIRS = {
    "contratos_vigentes": CONTRATOS_VIGENTES_DIR,
    "pedidos": PEDIDOS_DIR,
    "contratos_negociacoes": CONTRATOS_NEGOCIACOES_DIR,
    "cep_ibge": CEP_IBGE_DIR,
    "regioes_logisticas": REGIOES_LOGISTICAS_DIR,
}


@dataclass(frozen=True)
class DatasetFile:
    kind: str
    path: Path
    source: str = "configured"

    def as_dict(self) -> dict[str, Any]:
        stat = self.path.stat()
        meta = inspect_file(self.path)
        return {
            "kind": self.kind,
            "name": self.path.name,
            "path": str(self.path),
            "source": self.source,
            "size": stat.st_size,
            "modified": stat.st_mtime,
            **meta,
        }


def ensure_directories() -> None:
    for directory in DATASET_DIRS.values():
        try:
            directory.mkdir(parents=True, exist_ok=True)
        except OSError:
            # Serverless providers can expose the source tree as read-only.
            # Missing dataset directories are handled later as empty datasets.
            pass


def clear_dataset_cache() -> None:
    _inspect_file_cached.cache_clear()
    _load_file_cached.cache_clear()


def _is_supported(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS and not path.name.startswith("~$")


def _is_sqlite(path: Path) -> bool:
    return path.suffix.lower() in {".sqlite", ".db"}


def _quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def list_dataset_files(kind: str) -> list[DatasetFile]:
    ensure_directories()
    directory = DATASET_DIRS[kind]
    files = [DatasetFile(kind, path) for path in sorted(directory.iterdir()) if _is_supported(path)] if directory.exists() else []

    if kind == "contratos_vigentes" and not files:
        legacy_files = [p for p in sorted(LEGACY_DATA_DIR.glob("*")) if _is_supported(p)] if LEGACY_DATA_DIR.exists() else []
        files.extend(DatasetFile(kind, path, "legacy_data") for path in legacy_files)

    return files


@lru_cache(maxsize=64)
def _inspect_file_cached(path_text: str, modified: float) -> dict[str, Any]:
    path = Path(path_text)
    try:
        if _is_sqlite(path):
            with sqlite3.connect(path) as conn:
                tables = pd.read_sql_query(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' AND substr(name, 1, 1) != '_' ORDER BY name",
                    conn,
                )["name"].tolist()
                if not tables:
                    return {"sheets": [], "rows": 0, "columns": [], "error": "Banco SQLite sem tabelas de dados."}
                table = str(tables[0])
                sample = pd.read_sql_query(f"SELECT * FROM {_quote_identifier(table)} LIMIT 5", conn)
                row_count = int(
                    pd.read_sql_query(f"SELECT COUNT(*) AS total FROM {_quote_identifier(table)}", conn)["total"].iloc[0]
                )
                return {
                    "sheets": [str(item) for item in tables],
                    "activeSheet": table,
                    "rows": row_count,
                    "columns": list(sample.columns),
                    "error": None,
                    "format": "sqlite",
                }
        if path.suffix.lower() == ".csv":
            sample = pd.read_csv(path, nrows=5)
            return {"sheets": ["CSV"], "rows": None, "columns": list(sample.columns), "error": None}
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        columns = [str(col).strip() if col is not None else "" for col in header_row]
        return {
            "sheets": wb.sheetnames,
            "activeSheet": wb.sheetnames[0],
            "rows": max(sheet.max_row - 1, 0),
            "columns": columns,
            "error": None,
        }
    except Exception as exc:
        return {"sheets": [], "rows": 0, "columns": [], "error": str(exc)}


def inspect_file(path: Path) -> dict[str, Any]:
    return dict(_inspect_file_cached(str(path), path.stat().st_mtime))


@lru_cache(maxsize=32)
def _load_file_cached(path_text: str, sheet: str | None, modified: float) -> pd.DataFrame:
    path = Path(path_text)
    if _is_sqlite(path):
        with sqlite3.connect(path) as conn:
            table = sheet
            if not table:
                tables = pd.read_sql_query(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' AND substr(name, 1, 1) != '_' ORDER BY name",
                    conn,
                )["name"].tolist()
                if not tables:
                    return pd.DataFrame()
                table = str(tables[0])
            return clean_dataframe(pd.read_sql_query(f"SELECT * FROM {_quote_identifier(str(table))}", conn))
    if path.suffix.lower() == ".csv":
        return clean_dataframe(pd.read_csv(path))
    df = pd.read_excel(path, sheet_name=sheet or 0, dtype=object)
    return clean_dataframe(df)


def load_file(path: Path, sheet: str | None = None) -> pd.DataFrame:
    return _load_file_cached(str(path), sheet, path.stat().st_mtime).copy()


def load_dataset(kind: str) -> tuple[pd.DataFrame, list[dict[str, Any]], list[str]]:
    if db.database_configured():
        df, file_infos, errors = dataset_repository.active_dataset(kind)
        if not df.empty or file_infos:
            return df, file_infos, errors
        if db.is_vercel_runtime():
            return df, file_infos, [f"Nenhuma versao ativa da base {kind} encontrada no Supabase."]

    frames: list[pd.DataFrame] = []
    file_infos: list[dict[str, Any]] = []
    errors: list[str] = []

    for dataset_file in list_dataset_files(kind):
        info = dataset_file.as_dict()
        file_infos.append(info)
        if info.get("error"):
            errors.append(f"{dataset_file.path.name}: {info['error']}")
            continue
        try:
            df = load_file(dataset_file.path, info.get("activeSheet"))
            df["_arquivo_origem"] = dataset_file.path.name
            df["_aba_origem"] = info.get("activeSheet") or ""
            frames.append(df)
        except Exception as exc:
            errors.append(f"{dataset_file.path.name}: {exc}")

    if not frames:
        return pd.DataFrame(), file_infos, errors
    return pd.concat(frames, ignore_index=True, sort=False), file_infos, errors


def datasets_status() -> dict[str, Any]:
    if db.database_configured():
        db_status = dataset_repository.datasets_status(list(DATASET_DIRS.keys()))
        return {
            kind: {
                "directory": "supabase:dataset_versions",
                "source": "supabase",
                "files": db_status.get(kind, []),
            }
            for kind in DATASET_DIRS
        }

    return {
        kind: {
            "directory": str(DATASET_DIRS[kind]),
            "source": "local",
            "files": [item.as_dict() for item in list_dataset_files(kind)],
        }
        for kind in DATASET_DIRS
    }
