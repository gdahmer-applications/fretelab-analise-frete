"""Geração de arquivos (XLSX e PDF) a partir do payload do front-end.

O app roda no navegador (cálculo em JS) e envia o resultado consolidado para
o backend Flask apenas para exportação.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _safe_list(x: Any) -> List[Any]:
    return list(x) if isinstance(x, (list, tuple)) else []


MAX_WEIGHTS = 200
MAX_ROWS = 2000
MAX_VARIATIONS = 2000
MAX_CALC_ROWS = 200_000


def _is_number(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _validate_export_payload(payload: Dict[str, Any]) -> Tuple[List[int], List[Any], List[Any], List[Any], List[Any]]:
    if not isinstance(payload, dict):
        raise ValueError("Payload invalido.")

    weights_raw = _safe_list(payload.get("weights"))
    rows = _safe_list(payload.get("rows"))
    variations = _safe_list(payload.get("variations"))
    rep = _safe_list(payload.get("rep"))
    calc = _safe_list(payload.get("calc"))

    if not weights_raw:
        raise ValueError("Payload sem pesos para exportacao.")
    if len(weights_raw) > MAX_WEIGHTS:
        raise ValueError(f"Quantidade de pesos acima do limite ({MAX_WEIGHTS}).")

    weights: List[int] = []
    for w in weights_raw:
        if not _is_number(w):
            raise ValueError("Todos os pesos devem ser numericos.")
        wi = int(w)
        if wi <= 0:
            raise ValueError("Todos os pesos devem ser maiores que zero.")
        weights.append(wi)

    if not rows:
        raise ValueError("Payload sem linhas de analise para exportacao.")
    if len(rows) > MAX_ROWS:
        raise ValueError(f"Quantidade de linhas de analise acima do limite ({MAX_ROWS}).")
    if len(variations) > MAX_VARIATIONS:
        raise ValueError(f"Quantidade de variacoes acima do limite ({MAX_VARIATIONS}).")
    if len(calc) > MAX_CALC_ROWS:
        raise ValueError(f"Quantidade de linhas de calculo acima do limite ({MAX_CALC_ROWS}).")

    return weights, rows, variations, rep, calc


def build_xlsx_bytes(payload: Dict[str, Any]) -> bytes:
    """Cria um XLSX com duas abas: Analise e Calculos."""
    weights, rows, variations, rep, calc = _validate_export_payload(payload)

    wb = Workbook()

    # -------------------------
    # Aba: Analise
    # -------------------------
    ws = wb.active
    ws.title = "Analise"

    title = str(payload.get("title") or "RESULTADO ANÁLISE")
    meta = str(payload.get("meta") or "")
    filters = payload.get("filters") or {}

    # Cabeçalho
    ws.append([title])
    ws.append([meta])
    ws.append([
        f"Cidade: {filters.get('city','') or '—'} | UF: {filters.get('uf','') or '—'} | CEP: {filters.get('cep','') or '—'} | Estoques: {', '.join(filters.get('estoques') or []) or '—'}"
    ])
    ws.append([])

    # Tabela principal
    ws.append([""] + weights)

    money_fmt = "0.00"
    pct_fmt = "0%"

    # Linhas de cenário
    for r in rows:
        label = str(r.get("label") or "")
        totals = _safe_list(r.get("totals"))
        ws.append([label] + [v if isinstance(v, (int, float)) else None for v in totals])

    # Linhas de variação
    if variations:
        ws.append([])
        for vrow in variations:
            label = str(vrow.get("label") or "")
            vals = _safe_list(vrow.get("values"))
            ws.append([label] + [v if isinstance(v, (int, float)) else None for v in vals])

    # Representatividade
    if rep:
        ws.append([])
        ws.append(["REPRESENTATIVIDADE"] + [v if isinstance(v, (int, float)) else None for v in rep])

    # Estilos simples
    bold = Font(bold=True)
    ws[1][0].font = Font(bold=True, size=14)
    ws[5][0].font = bold
    for c in ws[5][1:]:
        c.font = bold
        c.alignment = Alignment(horizontal="center")

    # Formatação numérica: valores em dinheiro na tabela principal
    start_row = 6
    end_row = start_row + len(rows) - 1
    if len(rows) > 0:
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=2, max_col=1 + len(weights)):
            for cell in row:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="center")

    # Formatação percent nas variações e rep
    # Variações começam após (6 + len(rows) + 1) se existirem
    if variations:
        var_start = 6 + len(rows) + 2
        var_end = var_start + len(variations) - 1
        for row in ws.iter_rows(min_row=var_start, max_row=var_end, min_col=2, max_col=1 + len(weights)):
            for cell in row:
                cell.number_format = pct_fmt
                cell.alignment = Alignment(horizontal="center")

    if rep:
        rep_row = ws.max_row
        for cell in ws.iter_rows(min_row=rep_row, max_row=rep_row, min_col=2, max_col=1 + len(weights)):
            for c in cell:
                c.number_format = pct_fmt
                c.alignment = Alignment(horizontal="center")

    # Ajuste de largura
    ws.column_dimensions["A"].width = 38
    for idx in range(2, 2 + len(weights)):
        ws.column_dimensions[ws.cell(row=5, column=idx).column_letter].width = 12

    # -------------------------
    # Aba: Calculos
    # -------------------------
    ws2 = wb.create_sheet("Calculos")
    headers = [
        "Dataset",
        "Cenário",
        "Transportadora",
        "Peso",
        "Faixa KG",
        "Base",
        "Fixos",
        "% Nota",
        "Min aplicado",
        "Total",
    ]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = bold

    for r in calc:
        ws2.append([
            str(r.get("dataset") or ""),
            str(r.get("label") or ""),
            str(r.get("carrier") or ""),
            r.get("weight") if isinstance(r.get("weight"), (int, float)) else None,
            r.get("chosenKg") if isinstance(r.get("chosenKg"), (int, float)) else None,
            r.get("base") if isinstance(r.get("base"), (int, float)) else None,
            r.get("fixed") if isinstance(r.get("fixed"), (int, float)) else None,
            r.get("perc") if isinstance(r.get("perc"), (int, float)) else None,
            r.get("minApplied") if isinstance(r.get("minApplied"), (int, float)) else None,
            r.get("total") if isinstance(r.get("total"), (int, float)) else None,
        ])

    # Formatação numérica
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "0"

    for row in ws2.iter_rows(min_row=2, min_col=6, max_col=10):
        for cell in row:
            cell.number_format = money_fmt

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 32
    for col in ["D", "E"]:
        ws2.column_dimensions[col].width = 10
    for col in ["F", "G", "H", "I", "J"]:
        ws2.column_dimensions[col].width = 14

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_pdf_bytes(payload: Dict[str, Any]) -> bytes:
    """Gera um PDF (resumo da análise)."""
    # Import tardio para manter import mais leve.
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    weights, rows, variations, rep, _calc = _validate_export_payload(payload)
    filters = payload.get("filters") or {}

    title = str(payload.get("title") or "RESULTADO ANÁLISE")
    meta = str(payload.get("meta") or "")

    def fmt_money(v: Optional[float]) -> str:
        if v is None:
            return "—"
        try:
            return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return "—"

    def fmt_pct(v: Optional[float]) -> str:
        if v is None:
            return "—"
        try:
            return f"{float(v)*100:.0f}%".replace(".", ",")
        except Exception:
            return "—"

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    story.append(Spacer(1, 6))
    if meta:
        story.append(Paragraph(meta, styles["Normal"]))
        story.append(Spacer(1, 4))

    filt_line = f"Cidade: {filters.get('city','') or '—'} | UF: {filters.get('uf','') or '—'} | CEP: {filters.get('cep','') or '—'} | Estoques: {', '.join(filters.get('estoques') or []) or '—'}"
    story.append(Paragraph(filt_line, styles["Normal"]))
    story.append(Spacer(1, 10))

    # Monta tabela
    header = [""] + [str(int(w)) for w in weights]
    data: List[List[str]] = [header]

    for r in rows:
        label = str(r.get("label") or "")
        totals = _safe_list(r.get("totals"))
        data.append([label] + [fmt_money(v if isinstance(v, (int, float)) else None) for v in totals])

    if variations:
        data.append(["" for _ in header])
        for vrow in variations:
            label = str(vrow.get("label") or "")
            vals = _safe_list(vrow.get("values"))
            data.append([label] + [fmt_pct(v if isinstance(v, (int, float)) else None) for v in vals])

    if rep:
        data.append(["" for _ in header])
        data.append(["REPRESENTATIVIDADE"] + [fmt_pct(v if isinstance(v, (int, float)) else None) for v in rep])

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ]
        )
    )
    story.append(tbl)
    doc.build(story)
    return buf.getvalue()
