from __future__ import annotations

from datetime import datetime
import gc
from io import BytesIO
import hmac
import logging
import os
from pathlib import Path
import secrets
import sqlite3
import tempfile

from flask import Flask, abort, request, send_file, send_from_directory, session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from core.analysis_engine import available_carriers, available_carriers_by_location, generate_analysis, location_options, resolve_location_from_cep
from core import blob_storage, dataset_repository, db
from core.exporter import build_analysis_pdf, build_final_html, build_final_pdf
from core.history import delete_analysis, list_analyses, load_analysis, set_archived
from core.importer import load_xlsx_dataset
from core.insights import build_dashboard_summary
from core.logistics import preview_logistics
from core.normalization import dataframe_sample, norm_key
from core.repository import DATASET_DIRS, clear_dataset_cache, datasets_status, ensure_directories, load_dataset, load_file
from core.settings import BASE_DIR, LOGS_DIR, STATIC_DIR, SUPPORTED_EXTENSIONS, WEB_DIR
from core.validation import resolve_contract_columns, validate_contracts, validate_pedidos
from utils.export_files import build_pdf_bytes, build_xlsx_bytes

ensure_directories()
if os.getenv("VERCEL") != "1":
    LOGS_DIR.mkdir(parents=True, exist_ok=True)


def load_local_env() -> None:
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            os.environ.setdefault(key, value)


load_local_env()

app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path="/static")
app.config["MAX_CONTENT_LENGTH"] = int(os.getenv("MAX_UPLOAD_MB", "30")) * 1024 * 1024
app.secret_key = os.getenv("FRETELAB_SECRET_KEY") or secrets.token_hex(32)

ADMIN_PASSWORD = os.getenv("FRETELAB_ADMIN_PASSWORD", "")
DATASET_SQLITE_NAMES = {
    "contratos_vigentes": "dados.sqlite",
    "contratos_negociacoes": "negociacoes.sqlite",
    "pedidos": "pedidos.sqlite",
    "cep_ibge": "cep_ibge.sqlite",
    "regioes_logisticas": "regioes_logisticas.sqlite",
}
CONTRACT_TEMPLATE_PATH = BASE_DIR / "input" / "templates" / "template_contratos_fretelab.xlsx"

if os.getenv("VERCEL") == "1":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
else:
    logging.basicConfig(
        filename=str(LOGS_DIR / "frete_app.log"),
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )


def json_error(message: str, status: int = 400):
    return {"error": message}, status


def persistence_error(exc: Exception):
    if isinstance(exc, (db.PersistenceNotConfigured, blob_storage.BlobNotConfigured)):
        return json_error(str(exc), 503)
    return json_error(str(exc), 400)


def is_admin() -> bool:
    return bool(session.get("is_admin"))


def require_admin():
    if not is_admin():
        return json_error("Modo ADM necessario para esta operacao.", 403)
    return None


def require_dataset_persistence() -> bool:
    if db.database_configured():
        if not blob_storage.blob_configured():
            raise blob_storage.BlobNotConfigured("BLOB_READ_WRITE_TOKEN nao configurado para persistir arquivos.")
        return True
    if db.is_vercel_runtime():
        raise db.PersistenceNotConfigured("DATABASE_URL nao configurado; manutencao ADM persistente indisponivel.")
    return False


def tmp_workspace(kind: str, timestamp: str) -> Path:
    root = Path(tempfile.gettempdir()) / "fretelab" / kind / timestamp
    root.mkdir(parents=True, exist_ok=True)
    return root


def dataset_directory(kind: str) -> Path:
    if kind not in DATASET_DIRS:
        raise KeyError(kind)
    return DATASET_DIRS[kind]


def resolve_dataset_file(kind: str, filename: str) -> Path:
    directory = dataset_directory(kind).resolve()
    path = (directory / Path(filename).name).resolve()
    if path.parent != directory:
        raise ValueError("Nome de arquivo invalido.")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("Extensao de arquivo nao permitida.")
    return path


def validate_sqlite_file(path: Path) -> None:
    conn = None
    try:
        conn = sqlite3.connect(path)
        cursor = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' AND substr(name, 1, 1) != '_'"
        )
        tables = cursor.fetchall()
        cursor.close()
    except sqlite3.Error as exc:
        raise ValueError(f"Arquivo SQLite invalido: {exc}") from exc
    finally:
        if conn is not None:
            conn.close()
    if not tables:
        raise ValueError("Arquivo SQLite sem tabelas de dados.")


def write_dataset_sqlite(df: pd.DataFrame, target: Path, kind: str, source_file: str = "") -> None:
    target_table = kind
    conn = sqlite3.connect(target)
    try:
        df.to_sql(target_table, conn, if_exists="replace", index=False)
        conn.execute("CREATE TABLE IF NOT EXISTS _metadata (chave TEXT PRIMARY KEY, valor TEXT)")
        conn.execute("DELETE FROM _metadata")
        conn.executemany(
            "INSERT INTO _metadata (chave, valor) VALUES (?, ?)",
            [
                ("source_file", source_file),
                ("source_sheet", "0"),
                ("rows", str(len(df))),
                ("columns", str(len(df.columns))),
                ("imported_at", datetime.now().isoformat(timespec="seconds")),
            ],
        )
        conn.commit()
    finally:
        conn.close()


def convert_xlsx_to_sqlite(source: Path, target: Path, kind: str) -> None:
    safe_unlink(target)
    df = load_xlsx_dataset(source, kind)
    write_dataset_sqlite(df, target, kind, source.name)


def upload_and_register_file(
    path: Path,
    *,
    purpose: str,
    dataset_kind: str | None,
    original_filename: str,
    content_type: str,
    blob_folder: str,
) -> str:
    pathname = blob_storage.blob_path(blob_folder, blob_storage.safe_blob_name(path.name))
    blob_info = blob_storage.upload_path(path, pathname, content_type=content_type).as_dict()
    return dataset_repository.register_uploaded_file(
        purpose=purpose,
        dataset_kind=dataset_kind,
        original_filename=original_filename,
        content_type=content_type,
        size_bytes=path.stat().st_size,
        blob_info=blob_info,
    )


def persist_dataset_upload(kind: str, upload, *, mode: str, current_df: pd.DataFrame | None = None) -> dict[str, object]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    workspace = tmp_workspace(kind, timestamp)
    safe_name = blob_storage.safe_blob_name(upload.filename or f"{kind}.xlsx")
    upload_temp = workspace / safe_name
    sqlite_temp = workspace / DATASET_SQLITE_NAMES[kind]
    upload.save(upload_temp)

    new_df = load_xlsx_dataset(upload_temp, kind)
    if current_df is not None and not current_df.empty:
        final_df = pd.concat([current_df, new_df], ignore_index=True, sort=False)
    else:
        final_df = new_df

    write_dataset_sqlite(final_df, sqlite_temp, kind, safe_name)
    validate_sqlite_file(sqlite_temp)

    source_file_id = upload_and_register_file(
        upload_temp,
        purpose="dataset_original",
        dataset_kind=kind,
        original_filename=safe_name,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        blob_folder=blob_storage.blob_path("datasets", "original", kind),
    )
    sqlite_file_id = upload_and_register_file(
        sqlite_temp,
        purpose="dataset_sqlite_backup",
        dataset_kind=kind,
        original_filename=sqlite_temp.name,
        content_type="application/vnd.sqlite3",
        blob_folder=blob_storage.blob_path("datasets", "sqlite", kind),
    )
    version_id = dataset_repository.create_dataset_version(
        kind=kind,
        df=final_df,
        source_file_id=source_file_id,
        sqlite_file_id=sqlite_file_id,
        version_label=f"{kind}_{timestamp}",
        metadata={"mode": mode, "source_file": safe_name, "rows_added": int(len(new_df))},
    )
    dataset_repository.audit(
        f"dataset.{mode}",
        "dataset_version",
        version_id,
        {"kind": kind, "sourceFileId": source_file_id, "sqliteFileId": sqlite_file_id, "rows": int(len(final_df))},
        request.remote_addr,
    )
    clear_dataset_cache()
    return {
        "ok": True,
        "kind": kind,
        "file": DATASET_SQLITE_NAMES[kind],
        "source": safe_name,
        "versionId": version_id,
        "addedRows": int(len(new_df)),
        "totalRows": int(len(final_df)),
    }


def persist_dataset_dataframe(kind: str, df: pd.DataFrame, *, mode: str, metadata: dict[str, object]) -> dict[str, object]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    workspace = tmp_workspace(kind, timestamp)
    sqlite_temp = workspace / DATASET_SQLITE_NAMES[kind]
    write_dataset_sqlite(df, sqlite_temp, kind, mode)
    validate_sqlite_file(sqlite_temp)
    sqlite_file_id = upload_and_register_file(
        sqlite_temp,
        purpose="dataset_sqlite_backup",
        dataset_kind=kind,
        original_filename=sqlite_temp.name,
        content_type="application/vnd.sqlite3",
        blob_folder=blob_storage.blob_path("datasets", "sqlite", kind),
    )
    version_id = dataset_repository.create_dataset_version(
        kind=kind,
        df=df,
        source_file_id=None,
        sqlite_file_id=sqlite_file_id,
        version_label=f"{kind}_{timestamp}",
        metadata={"mode": mode, **metadata},
    )
    dataset_repository.audit(
        f"dataset.{mode}",
        "dataset_version",
        version_id,
        {"kind": kind, "sqliteFileId": sqlite_file_id, **metadata},
        request.remote_addr,
    )
    clear_dataset_cache()
    return {"ok": True, "kind": kind, "file": DATASET_SQLITE_NAMES[kind], "versionId": version_id}


def save_export_to_blob(name: str, content: bytes, content_type: str, analysis_id: str | None = None) -> None:
    if not (db.database_configured() and blob_storage.blob_configured()):
        return
    pathname = blob_storage.blob_path("exports", analysis_id or "session", blob_storage.safe_blob_name(name))
    blob_info = blob_storage.upload_bytes(pathname, content, content_type=content_type).as_dict()
    dataset_repository.register_uploaded_file(
        purpose="analysis_export",
        dataset_kind=None,
        original_filename=name,
        content_type=content_type,
        size_bytes=len(content),
        blob_info=blob_info,
    )


def safe_unlink(path: Path) -> None:
    try:
        if path.exists():
            path.unlink()
    except PermissionError:
        app.logger.warning("Nao foi possivel remover arquivo temporario bloqueado: %s", path)


def remove_supported_dataset_files(kind: str) -> None:
    directory = dataset_directory(kind)
    for path in directory.iterdir():
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS and not path.name.startswith(("~$", ".")):
            path.unlink()


def is_weight_column(column: object) -> bool:
    text = str(column).strip()
    if not text:
        return False
    try:
        value = float(text.replace(",", "."))
    except ValueError:
        return False
    return value > 0 and abs(value - round(value)) < 0.0001


def display_header(column: object) -> str:
    text = str(column).strip()
    normalized = {
        "FRETE TOTAL MINIMO": "FRETE TOTAL M\u00cdNIMO",
        "PEDAGIO VALOR FIXO": "PED\u00c1GIO VALOR FIXO",
        "PEDAGIO FRACAO A CADA X KG": "PED\u00c1GIO FRA\u00c7\u00c3O A CADA x KG",
        "GRIS MINIMO": "GRIS M\u00cdNIMO",
        "SEGURO MINIMO": "SEGURO M\u00cdNIMO",
        "REGIAO LOGISTICA": "REGI\u00c3O LOG\u00cdSTICA",
    }
    return normalized.get(norm_key(text), text)


def export_contract_dataset_xlsx(df: pd.DataFrame, sheet_name: str) -> BytesIO:
    metadata_cols = {col for col in df.columns if str(col).startswith("_")}
    id_col = "ID" if "ID" in df.columns else "ID INTELIPOST" if "ID INTELIPOST" in df.columns else None
    cadastro_order = [
        id_col,
        "NOME",
        "CNPJ",
        "ESTOQUE",
        "CIDADE",
        "UF",
        "REGIAO LOGISTICA",
        "REGIÃO LOGÍSTICA",
        "CEPI",
        "CEPF",
        "PRAZO(DIAS ÚTEIS)",
        "PRAZO(DIAS UTEIS)",
    ]
    cadastro = []
    for col in cadastro_order:
        if col and col in df.columns and col not in cadastro:
            cadastro.append(col)

    weight_cols = sorted(
        [col for col in df.columns if col not in metadata_cols and is_weight_column(col)],
        key=lambda value: int(float(str(value).replace(",", "."))),
    )
    generalidades = [
        col for col in df.columns
        if col not in metadata_cols and col not in cadastro and col not in weight_cols
    ]
    columns = cadastro + generalidades + weight_cols

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    groups = ["CADASTRO"] * len(cadastro) + ["GENERALIDADES"] * len(generalidades) + ["FRETE PESO"] * len(weight_cols)
    ws.append(groups)
    ws.append([display_header(col) for col in columns])
    for row in df[columns].itertuples(index=False, name=None):
        ws.append(list(row))

    fills = {
        "CADASTRO": "1F4E78",
        "GENERALIDADES": "7F6000",
        "FRETE PESO": "375623",
    }
    thin = Side(style="thin", color="2B313D")
    for col_idx, group in enumerate(groups, start=1):
        for row_idx in (1, 2):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = PatternFill("solid", fgColor=fills[group])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        ws.column_dimensions[get_column_letter(col_idx)].width = 16 if group != "FRETE PESO" else 11

    start = 1
    while start <= len(groups):
        group = groups[start - 1]
        end = start
        while end <= len(groups) and groups[end - 1] == group:
            end += 1
        if end - start > 1:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end - 1)
        start = end

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(columns))}{max(ws.max_row, 2)}"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_dataset_xlsx(df: pd.DataFrame, kind: str) -> BytesIO:
    if kind in {"contratos_vigentes", "contratos_negociacoes"}:
        return export_contract_dataset_xlsx(df, kind)
    content = BytesIO()
    clean_columns = {col: display_header(col) for col in df.columns}
    with pd.ExcelWriter(content, engine="openpyxl") as writer:
        df.rename(columns=clean_columns).to_excel(writer, index=False, sheet_name=kind[:31])
    content.seek(0)
    return content


@app.get("/api/admin/status")
def admin_status():
    return {"isAdmin": is_admin(), "configured": bool(ADMIN_PASSWORD)}


@app.post("/api/admin/login")
def admin_login():
    if not ADMIN_PASSWORD:
        return json_error("Defina a variavel FRETELAB_ADMIN_PASSWORD no servidor para habilitar o modo ADM.", 503)
    payload = request.get_json(silent=True) or {}
    password = str(payload.get("password") or "")
    if not hmac.compare_digest(password, ADMIN_PASSWORD):
        return json_error("Senha ADM invalida.", 401)
    session["is_admin"] = True
    return {"isAdmin": True}


@app.post("/api/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return {"isAdmin": False}


@app.get("/api/admin/datasets/<kind>/files/<filename>/download")
def admin_download_dataset(kind: str, filename: str):
    blocked = require_admin()
    if blocked:
        return blocked
    if db.database_configured():
        df, files, errors = load_dataset(kind)
        if df.empty:
            return json_error("; ".join(errors) or "Base nao encontrada.", 404)
        content = export_dataset_xlsx(df, kind)
        return send_file(
            content,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{Path(filename).stem or kind}.xlsx",
            max_age=0,
        )
    try:
        path = resolve_dataset_file(kind, filename)
    except (KeyError, ValueError) as exc:
        return json_error(str(exc), 400)
    if not path.exists():
        return json_error("Arquivo nao encontrado.", 404)
    try:
        info_df = load_file(path)
        content = export_dataset_xlsx(info_df, kind)
        return send_file(
            content,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{Path(filename).stem}.xlsx",
            max_age=0,
        )
    except Exception as exc:
        return json_error(f"Falha ao exportar base em XLSX: {exc}", 400)


@app.get("/api/templates/contratos")
def download_contract_template():
    if not CONTRACT_TEMPLATE_PATH.exists():
        return json_error("Template de contratos nao encontrado.", 404)
    return send_file(
        CONTRACT_TEMPLATE_PATH,
        as_attachment=True,
        download_name=CONTRACT_TEMPLATE_PATH.name,
        max_age=0,
    )


@app.post("/api/admin/datasets/<kind>/replace")
def admin_replace_dataset(kind: str):
    blocked = require_admin()
    if blocked:
        return blocked
    if kind not in DATASET_SQLITE_NAMES:
        return json_error("Tipo de base invalido.", 400)
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return json_error("Envie um arquivo .xlsx.", 400)
    suffix = Path(upload.filename).suffix.lower()
    if suffix != ".xlsx":
        return json_error("O replace pelo portal recebe apenas .xlsx e converte automaticamente para SQLite.", 400)
    try:
        if require_dataset_persistence():
            result = persist_dataset_upload(kind, upload, mode="replace")
            app.logger.info("Base %s substituida no Supabase/Blob por %s", kind, result.get("source"))
            return result
    except Exception as exc:
        return persistence_error(exc)

    directory = dataset_directory(kind)
    directory.mkdir(parents=True, exist_ok=True)
    target = directory / DATASET_SQLITE_NAMES[kind]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    upload_temp = directory / f".{target.stem}_{timestamp}.upload.xlsx"
    sqlite_temp = directory / f".{target.stem}_{timestamp}.upload.sqlite"
    try:
        upload.save(upload_temp)
        convert_xlsx_to_sqlite(upload_temp, sqlite_temp, kind)
        validate_sqlite_file(sqlite_temp)
        gc.collect()
        archive_dir = directory / "origem"
        archive_dir.mkdir(exist_ok=True)
        archived_source = archive_dir / f"{target.stem}_{timestamp}.xlsx"
        upload_temp.replace(archived_source)
        remove_supported_dataset_files(kind)
        sqlite_temp.replace(target)
        clear_dataset_cache()
        app.logger.info("Base %s substituida por %s a partir de %s", kind, target.name, archived_source.name)
        return {"ok": True, "kind": kind, "file": target.name, "source": archived_source.name}
    except Exception as exc:
        safe_unlink(upload_temp)
        safe_unlink(sqlite_temp)
        return json_error(str(exc), 400)


@app.get("/api/admin/datasets/contratos_negociacoes/carriers")
def admin_negotiation_carriers():
    blocked = require_admin()
    if blocked:
        return blocked
    df, _files, errors = load_dataset("contratos_negociacoes")
    if df.empty:
        return {"carriers": [], "loadErrors": errors}
    name_col = resolve_contract_columns(df).get("nome")
    if not name_col:
        return {"carriers": [], "loadErrors": errors + ["Coluna de transportadora nao localizada."]}
    carriers = sorted({str(value).strip() for value in df[name_col].dropna().tolist() if str(value).strip()})
    return {"carriers": carriers, "loadErrors": errors}


@app.post("/api/admin/datasets/contratos_negociacoes/append")
def admin_append_negotiation_dataset():
    blocked = require_admin()
    if blocked:
        return blocked
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return json_error("Envie um arquivo .xlsx.", 400)
    if Path(upload.filename).suffix.lower() != ".xlsx":
        return json_error("A inclusao por transportadora recebe apenas .xlsx.", 400)

    kind = "contratos_negociacoes"
    try:
        if require_dataset_persistence():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            workspace = tmp_workspace(kind, timestamp)
            safe_name = blob_storage.safe_blob_name(upload.filename)
            upload_temp = workspace / safe_name
            upload.save(upload_temp)
            new_df = load_xlsx_dataset(upload_temp, kind)
            validation = validate_contracts(new_df)
            if not validation.get("ok"):
                safe_unlink(upload_temp)
                return json_error("Campos faltantes no arquivo enviado: " + ", ".join(validation.get("missing", [])), 400)
            current_df, _files, _errors = load_dataset(kind)
            upload.stream.seek(0)
            result = persist_dataset_upload(kind, upload, mode="append", current_df=current_df)
            app.logger.info("Negociacao adicionada no Supabase/Blob: %s", result.get("source"))
            return result
    except Exception as exc:
        return persistence_error(exc)

    directory = dataset_directory(kind)
    directory.mkdir(parents=True, exist_ok=True)
    target = directory / DATASET_SQLITE_NAMES[kind]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    upload_temp = directory / f".{target.stem}_{timestamp}.append.xlsx"
    sqlite_temp = directory / f".{target.stem}_{timestamp}.append.sqlite"
    try:
        upload.save(upload_temp)
        new_df = load_xlsx_dataset(upload_temp, kind)
        validation = validate_contracts(new_df)
        if not validation.get("ok"):
            safe_unlink(upload_temp)
            return json_error("Campos faltantes no arquivo enviado: " + ", ".join(validation.get("missing", [])), 400)

        current_df = load_file(target) if target.exists() else pd.DataFrame()
        combined = pd.concat([current_df, new_df], ignore_index=True, sort=False) if not current_df.empty else new_df
        write_dataset_sqlite(combined, sqlite_temp, kind, upload.filename)
        validate_sqlite_file(sqlite_temp)

        archive_dir = directory / "origem"
        archive_dir.mkdir(exist_ok=True)
        archived_source = archive_dir / f"{target.stem}_append_{timestamp}.xlsx"
        upload_temp.replace(archived_source)
        sqlite_temp.replace(target)
        clear_dataset_cache()
        app.logger.info("Negociacao adicionada em %s a partir de %s", target.name, archived_source.name)
        return {"ok": True, "kind": kind, "file": target.name, "addedRows": int(len(new_df)), "totalRows": int(len(combined))}
    except Exception as exc:
        safe_unlink(upload_temp)
        safe_unlink(sqlite_temp)
        return json_error(str(exc), 400)


@app.post("/api/admin/datasets/contratos_negociacoes/delete-carrier")
def admin_delete_negotiation_carrier():
    blocked = require_admin()
    if blocked:
        return blocked
    payload = request.get_json(silent=True) or {}
    carrier = str(payload.get("carrier") or "").strip()
    if not carrier:
        return json_error("Informe a transportadora para remover.", 400)

    kind = "contratos_negociacoes"
    try:
        if require_dataset_persistence():
            df, _files, errors = load_dataset(kind)
            if df.empty:
                return json_error("; ".join(errors) or "Base de negociacoes nao encontrada.", 404)
            name_col = resolve_contract_columns(df).get("nome")
            if not name_col:
                return json_error("Coluna de transportadora nao localizada.", 400)
            wanted = norm_key(carrier)
            keep_mask = df[name_col].map(lambda value: norm_key(value) != wanted)
            removed = int((~keep_mask).sum())
            if removed == 0:
                return json_error("Transportadora nao encontrada na base de negociacoes.", 404)
            updated = df.loc[keep_mask].copy()
            result = persist_dataset_dataframe(
                kind,
                updated,
                mode="delete_carrier",
                metadata={"carrier": carrier, "removedRows": removed, "totalRows": int(len(updated))},
            )
            result.update({"carrier": carrier, "removedRows": removed, "totalRows": int(len(updated))})
            app.logger.info("Transportadora removida das negociacoes no Supabase: %s (%s linhas)", carrier, removed)
            return result
    except Exception as exc:
        return persistence_error(exc)

    directory = dataset_directory(kind)
    target = directory / DATASET_SQLITE_NAMES[kind]
    if not target.exists():
        return json_error("Base de negociacoes nao encontrada.", 404)
    sqlite_temp = directory / f".{target.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.delete.sqlite"
    try:
        df = load_file(target)
        name_col = resolve_contract_columns(df).get("nome")
        if not name_col:
            return json_error("Coluna de transportadora nao localizada.", 400)
        wanted = norm_key(carrier)
        keep_mask = df[name_col].map(lambda value: norm_key(value) != wanted)
        removed = int((~keep_mask).sum())
        if removed == 0:
            return json_error("Transportadora nao encontrada na base de negociacoes.", 404)
        updated = df.loc[keep_mask].copy()
        write_dataset_sqlite(updated, sqlite_temp, kind, f"delete:{carrier}")
        validate_sqlite_file(sqlite_temp)
        sqlite_temp.replace(target)
        clear_dataset_cache()
        app.logger.info("Transportadora removida das negociacoes: %s (%s linhas)", carrier, removed)
        return {"ok": True, "carrier": carrier, "removedRows": removed, "totalRows": int(len(updated))}
    except Exception as exc:
        safe_unlink(sqlite_temp)
        return json_error(str(exc), 400)


@app.delete("/api/admin/datasets/<kind>/files/<filename>")
def admin_delete_dataset_file(kind: str, filename: str):
    blocked = require_admin()
    if blocked:
        return blocked
    if db.database_configured():
        try:
            result = dataset_repository.delete_active_dataset(kind, request.remote_addr)
            clear_dataset_cache()
            app.logger.info("Base ativa removida no Supabase: %s", kind)
            return result
        except Exception as exc:
            return persistence_error(exc)
    if db.is_vercel_runtime():
        return json_error("DATABASE_URL nao configurado; exclusao persistente indisponivel.", 503)
    try:
        path = resolve_dataset_file(kind, filename)
    except (KeyError, ValueError) as exc:
        return json_error(str(exc), 400)
    if not path.exists():
        return json_error("Arquivo nao encontrado.", 404)
    path.unlink()
    clear_dataset_cache()
    app.logger.info("Arquivo de base removido: %s", path)
    return {"ok": True, "deleted": path.name}


@app.get("/")
def index():
    if not (WEB_DIR / "index.html").exists():
        abort(404)
    return send_from_directory(str(WEB_DIR), "index.html")


@app.get("/health")
def health():
    database = db.check_database()
    blob = blob_storage.check_blob()
    if database["configured"] and blob["configured"]:
        persistence_mode = "supabase_postgres+vercel_blob"
    elif db.is_vercel_runtime():
        persistence_mode = "unconfigured_serverless_read_only"
    else:
        persistence_mode = "local_development_files"
    return {
        "status": "ok",
        "baseDir": str(BASE_DIR),
        "databaseConfigured": database["configured"],
        "databaseOk": database["ok"],
        "databaseError": database["error"],
        "blobConfigured": blob["configured"],
        "blobOk": blob["ok"],
        "blobError": blob["error"],
        "persistenceMode": persistence_mode,
    }


@app.get("/api/files")
def files_status():
    status = datasets_status()
    validation = {}
    for kind in ("contratos_vigentes", "contratos_negociacoes", "pedidos"):
        df, _files, errors = load_dataset(kind)
        if kind == "pedidos":
            item = validate_pedidos(df) if not df.empty else {"ok": False, "missing": ["arquivo"], "columns": {}, "rowCount": 0}
        else:
            item = validate_contracts(df) if not df.empty else {"ok": kind == "contratos_negociacoes", "missing": ["arquivo"], "columns": {}, "rowCount": 0}
        item["loadErrors"] = errors
        validation[kind] = item
    return {"status": status, "validation": validation}


@app.get("/api/carriers")
def carriers():
    cep = request.args.get("cep", "")
    try:
        return available_carriers(cep)
    except Exception as exc:
        app.logger.exception("Falha ao buscar transportadoras: %s", exc)
        return json_error(str(exc), 400)


@app.get("/api/options")
def options():
    try:
        return location_options(
            uf=request.args.get("uf"),
            municipio=request.args.get("municipio"),
            logistics_region=request.args.get("logisticsRegion"),
        )
    except Exception as exc:
        app.logger.exception("Falha ao buscar opcoes: %s", exc)
        return json_error(str(exc), 400)


@app.get("/api/preview")
def preview_dataset():
    kind = request.args.get("kind", "contratos_vigentes")
    query = request.args.get("q", "").strip()
    try:
        limit = int(request.args.get("limit", "100") or 100)
    except (TypeError, ValueError):
        return json_error("Parametro limit invalido.", 400)
    limit = max(1, min(limit, 500))
    try:
        if kind == "regioes_logisticas":
            columns, rows = preview_logistics(limit=limit, query=query)
            return {"kind": kind, "columns": columns, "rows": rows, "rowCount": len(rows)}
        df, files, errors = load_dataset(kind)
        if df.empty:
            return {"kind": kind, "columns": [], "rows": [], "files": files, "errors": errors, "rowCount": 0}
        if query:
            import pandas as pd
            mask = pd.Series(False, index=df.index)
            for col in df.columns:
                mask = mask | df[col].astype(str).str.contains(query, case=False, na=False)
            df = df[mask]
        return {
            "kind": kind,
            "columns": [str(col) for col in df.columns],
            "rows": dataframe_sample(df, limit),
            "files": files,
            "errors": errors,
            "rowCount": int(len(df)),
        }
    except Exception as exc:
        app.logger.exception("Falha ao gerar preview: %s", exc)
        return json_error(str(exc), 400)


@app.post("/api/carriers/location")
def carriers_by_location():
    payload = request.get_json(silent=True) or {}
    try:
        return available_carriers_by_location(payload)
    except Exception as exc:
        app.logger.exception("Falha ao buscar transportadoras por localidade: %s", exc)
        return json_error(str(exc), 400)


@app.get("/api/cep/resolve")
def resolve_cep_location():
    try:
        return resolve_location_from_cep(request.args.get("cep", ""))
    except Exception as exc:
        app.logger.exception("Falha ao resolver CEP: %s", exc)
        return json_error(str(exc), 400)


@app.post("/api/analyses")
def create_analysis():
    payload = request.get_json(silent=True) or {}
    try:
        analysis = generate_analysis(payload)
        app.logger.info("Analise criada: %s", analysis.get("id"))
        return analysis, 201
    except Exception as exc:
        app.logger.exception("Falha ao gerar analise: %s", exc)
        return json_error(str(exc), 400)


@app.get("/api/analyses")
def analyses():
    archived_arg = request.args.get("archived")
    archived = None if archived_arg is None else archived_arg.lower() in {"1", "true", "yes"}
    return {"items": list_analyses(archived=archived)}


@app.get("/api/dashboard")
def dashboard():
    active_items = list_analyses(archived=False)
    archived_items = list_analyses(archived=True)
    records = []
    for item in active_items:
        record = load_analysis(item.get("id"))
        if record:
            records.append(record)
    return build_dashboard_summary(records, archived_count=len(archived_items))


@app.get("/api/analyses/<analysis_id>")
def analysis_detail(analysis_id: str):
    record = load_analysis(analysis_id)
    if not record:
        return json_error("Analise nao encontrada.", 404)
    return record


@app.post("/api/analyses/<analysis_id>/archive")
def archive_analysis(analysis_id: str):
    payload = request.get_json(silent=True) or {}
    archived = bool(payload.get("archived", True))
    try:
        return set_archived(analysis_id, archived)
    except FileNotFoundError as exc:
        return json_error(str(exc), 404)
    except Exception as exc:
        return persistence_error(exc)


@app.delete("/api/analyses/<analysis_id>")
def delete_analysis_route(analysis_id: str):
    try:
        return delete_analysis(analysis_id)
    except FileNotFoundError as exc:
        return json_error(str(exc), 404)
    except Exception as exc:
        return persistence_error(exc)


@app.get("/api/analyses/<analysis_id>/export/pdf")
def export_analysis_pdf(analysis_id: str):
    record = load_analysis(analysis_id)
    if not record:
        return json_error("Analise nao encontrada.", 404)
    content = build_analysis_pdf(record)
    filename = f"analise_frete_{analysis_id}.pdf"
    if request.args.get("save") in {"1", "true", "yes"}:
        save_export_to_blob(filename, content, "application/pdf", analysis_id)
    return send_file(BytesIO(content), mimetype="application/pdf", as_attachment=True, download_name=filename, max_age=0)


@app.get("/api/analyses/<analysis_id>/export/html")
def export_analysis_html(analysis_id: str):
    record = load_analysis(analysis_id)
    if not record:
        return json_error("Analise nao encontrada.", 404)
    content = build_final_html([record])
    filename = f"analise_frete_{analysis_id}.html"
    if request.args.get("save") in {"1", "true", "yes"}:
        save_export_to_blob(filename, content, "text/html", analysis_id)
    return send_file(BytesIO(content), mimetype="text/html", as_attachment=True, download_name=filename, max_age=0)


@app.get("/api/analyses/export/<ext>")
def export_final_html(ext: str):
    if ext.lower() not in {"htm", "html"}:
        return json_error("Extensao invalida para exportacao.", 400)
    records = []
    for item in reversed(list_analyses(archived=False)):
        record = load_analysis(item.get("id"))
        if record:
            records.append(record)
    content = build_final_html(records)
    filename = f"comparativo_transportadoras.{ext.lower()}"
    if request.args.get("save") in {"1", "true", "yes"}:
        save_export_to_blob(filename, content, "text/html")
    return send_file(BytesIO(content), mimetype="text/html", as_attachment=True, download_name=filename, max_age=0)


@app.post("/api/analyses/export/session/<ext>")
def export_session(ext: str):
    ext = ext.lower()
    if ext not in {"htm", "html", "pdf"}:
        return json_error("Extensao invalida para exportacao.", 400)
    payload = request.get_json(silent=True) or {}
    ids = payload.get("ids") or []
    records = []
    for analysis_id in ids:
        record = load_analysis(str(analysis_id))
        if record:
            records.append(record)
    if not records:
        return json_error("Nenhuma analise valida foi informada para exportacao.", 400)
    if ext == "pdf":
        content = build_final_pdf(records)
        if payload.get("save"):
            save_export_to_blob("comparativo_transportadoras_consolidado.pdf", content, "application/pdf")
        return send_file(
            BytesIO(content),
            mimetype="application/pdf",
            as_attachment=True,
            download_name="comparativo_transportadoras_consolidado.pdf",
            max_age=0,
        )
    content = build_final_html(records)
    if payload.get("save"):
        save_export_to_blob(f"comparativo_transportadoras_consolidado.{ext}", content, "text/html")
    return send_file(
        BytesIO(content),
        mimetype="text/html",
        as_attachment=True,
        download_name=f"comparativo_transportadoras_consolidado.{ext}",
        max_age=0,
    )


@app.post("/api/export/xlsx")
def export_xlsx_legacy():
    payload = request.get_json(silent=True)
    if payload is None:
        return json_error("Body JSON invalido.", 400)
    try:
        content = build_xlsx_bytes(payload)
    except Exception as exc:
        return json_error(str(exc), 400)
    return send_file(
        BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="resultado_analise.xlsx",
        max_age=0,
    )


@app.post("/api/export/pdf")
def export_pdf_legacy():
    payload = request.get_json(silent=True)
    if payload is None:
        return json_error("Body JSON invalido.", 400)
    try:
        content = build_pdf_bytes(payload)
    except Exception as exc:
        return json_error(str(exc), 400)
    return send_file(BytesIO(content), mimetype="application/pdf", as_attachment=True, download_name="resultado_analise.pdf", max_age=0)


@app.errorhandler(413)
def payload_too_large(_exc):
    return json_error("Payload muito grande.", 413)


if __name__ == "__main__":
    host = os.getenv("APP_HOST", "127.0.0.1")
    port = int(os.getenv("APP_PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "0").strip().lower() in {"1", "true", "yes"}
    app.run(host=host, port=port, debug=debug)
